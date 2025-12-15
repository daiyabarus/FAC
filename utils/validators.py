"""
Data Validators
Validate Excel files and data integrity
"""
import pandas as pd
from pathlib import Path
import openpyxl


def validate_excel_file(file_path: str, required_sheet: str = None) -> dict:
    """
    Validate Excel file exists and has required sheets

    Args:
        file_path: Path to Excel file
        required_sheet: Required sheet name (optional)

    Returns:
        Dictionary with validation results
    """
    result = {
        'valid': False,
        'error': None,
        'sheets': []
    }

    # Check file exists
    path = Path(file_path)
    if not path.exists():
        result['error'] = f"File not found: {file_path}"
        return result

    # Check file extension
    if path.suffix not in ['.xlsx', '.xls']:
        result['error'] = f"Invalid file type. Expected .xlsx or .xls, got {path.suffix}"
        return result

    try:
        # Load workbook to check sheets
        wb = openpyxl.load_workbook(file_path, read_only=True)
        result['sheets'] = wb.sheetnames

        # Check required sheet if specified
        if required_sheet and required_sheet not in wb.sheetnames:
            result['error'] = f"Required sheet '{required_sheet}' not found. Available sheets: {wb.sheetnames}"
            wb.close()
            return result

        wb.close()
        result['valid'] = True
        return result

    except Exception as e:
        result['error'] = f"Error reading Excel file: {str(e)}"
        return result


def validate_cluster_data(df: pd.DataFrame) -> dict:
    """
    Validate cluster data has required columns

    Args:
        df: Cluster DataFrame

    Returns:
        Dictionary with validation results
    """
    result = {
        'valid': False,
        'error': None,
        'missing_columns': []
    }

    required_columns = ['CLUSTER', 'TOWERID', 'LTE_CELL', 'TX', 'SITENAME']

    # Check DataFrame is not empty
    if df is None or len(df) == 0:
        result['error'] = "Cluster data is empty"
        return result

    # Check required columns by index (first 5 columns)
    if len(df.columns) < 5:
        result['error'] = f"Insufficient columns. Expected at least 5, got {len(df.columns)}"
        return result

    # Validate data types
    try:
        # Check for null values in critical columns
        critical_cols = [0, 1, 4]  # CLUSTER, TOWERID, SITENAME
        for col_idx in critical_cols:
            null_count = df.iloc[:, col_idx].isna().sum()
            if null_count > 0:
                result['error'] = f"Found {null_count} null values in column {col_idx}"
                return result

        result['valid'] = True
        return result

    except Exception as e:
        result['error'] = f"Error validating cluster data: {str(e)}"
        return result


def validate_lte_data(df: pd.DataFrame) -> dict:
    """
    Validate LTE data structure

    Args:
        df: LTE DataFrame

    Returns:
        Dictionary with validation results
    """
    result = {
        'valid': False,
        'error': None,
        'warnings': []
    }

    # Check minimum columns (61 expected)
    if len(df.columns) < 61:
        result['error'] = f"Insufficient columns for LTE data. Expected 61, got {len(df.columns)}"
        return result

    # Check for date column
    try:
        pd.to_datetime(df.iloc[:, 0])
    except:
        result['error'] = "First column (BEGIN_TIME) is not a valid date format"
        return result

    # Check for numeric columns
    numeric_start = 19  # LTE_RRC_SSR_NUM
    numeric_cols = df.iloc[:, numeric_start:]

    non_numeric = []
    for col in numeric_cols.columns[:10]:  # Check first 10 numeric columns
        if not pd.api.types.is_numeric_dtype(numeric_cols[col]):
            # Try to convert
            try:
                pd.to_numeric(numeric_cols[col], errors='coerce')
            except:
                non_numeric.append(col)

    if non_numeric:
        result['warnings'].append(
            f"Some columns may need numeric conversion: {non_numeric[:3]}")

    result['valid'] = True
    return result


def validate_gsm_data(df: pd.DataFrame) -> dict:
    """
    Validate GSM data structure

    Args:
        df: GSM DataFrame

    Returns:
        Dictionary with validation results
    """
    result = {
        'valid': False,
        'error': None,
        'warnings': []
    }

    # Check minimum columns (19 expected)
    if len(df.columns) < 19:
        result['error'] = f"Insufficient columns for GSM data. Expected 19, got {len(df.columns)}"
        return result

    # Check for date column
    try:
        pd.to_datetime(df.iloc[:, 0])
    except:
        result['error'] = "First column (BEGIN_TIME) is not a valid date format"
        return result

    # Check for numeric columns
    numeric_start = 13  # GSM_CSSR_NUM
    numeric_cols = df.iloc[:, numeric_start:]

    if len(numeric_cols.columns) < 6:
        result['error'] = "Missing numeric KPI columns in GSM data"
        return result

    result['valid'] = True
    return result


def validate_date_range(df: pd.DataFrame, date_col_idx: int = 0) -> dict:
    """
    Validate and get date range from DataFrame

    Args:
        df: DataFrame with date column
        date_col_idx: Index of date column

    Returns:
        Dictionary with date range info
    """
    result = {
        'valid': False,
        'start_date': None,
        'end_date': None,
        'num_days': 0,
        'error': None
    }

    try:
        dates = pd.to_datetime(df.iloc[:, date_col_idx])

        result['start_date'] = dates.min()
        result['end_date'] = dates.max()
        result['num_days'] = (dates.max() - dates.min()).days + 1
        result['valid'] = True

        return result

    except Exception as e:
        result['error'] = f"Error parsing dates: {str(e)}"
        return result
