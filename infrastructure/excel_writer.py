"""
Infrastructure Layer - Excel Report Writer (Template-Based)
File: infrastructure/excel_writer.py

Uses existing template 'datatemplate.xlsx', sheet 'Template', starting at A5.
Populates columns M-R (months 1-3: Value, Status pairs) for specific KPI rows.
"""

from pathlib import Path
from datetime import datetime
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

from domain.models import ClusterReport, KPIResult

class ExcelReportWriter:
    """Writes FAC KPI reports to Excel using existing template."""

    KPI_ROW_MAP = {
        ("Call Setup Success Rate", "2G RAN", "95.0", ">=", "98.5"): 10,
        ("SDCCH Success rate", "2G RAN", "95.0", ">=", "98.5"): 11,
        ("Perceive Drop Rate", "2G RAN", "95.0", "<", "2.0"): 12,
        ("Session Setup Success Rate", "4G RAN", "97.0", ">=", "99.0"): 13,
        ("RACH Success Rate", "4G RAN", "60.0", ">=", "85.0"): 14,
        ("RACH Success Rate", "4G RAN", "3.0", "<", "55.0"): 15,
        ("Handover Success Rate Inter and Intra-Frequency", "4G RAN", "95.0", ">=", "97.0"): 16,
        ("E-RAB Drop Rate", "4G RAN", "95.0", "<", "2.0"): 17,
        ("Downlink User Throughput", "4G RAN", "85.0", ">=", "3.0"): 18,
        ("Downlink User Throughput", "4G RAN", "2.0", "<", "1.0"): 19,
        ("Uplink User Throughput", "4G RAN", "65.0", ">=", "1.0"): 20,
        ("Uplink User Throughput", "4G RAN", "2.0", "<", "0.256"): 21,
        ("UL Packet Loss (PDCP )", "4G RAN", "97.0", "<", "0.85"): 22,
        ("DL Packet Loss (PDCP )", "4G RAN", "97.0", "<", "0.10"): 23,
        ("CQI", "4G RAN", "95.0", ">=", "7.0"): 24,
        ("MIMO Transmission Rank2 Rate", "4G RAN", "70.0", ">=", "35.0"): 25,
        ("MIMO Transmission Rank2 Rate", "4G RAN", "5.0", "<", "20.0"): 26,
        ("UL RSSI", "4G RAN", "97.0", "<", "-105.0"): 27,
        ("Packet Latency", "4G RAN", "95.0", "<=", "30.0"): 28,
        ("Packet Latency", "4G RAN", "5.0", ">", "40.0"): 29,
        ("Spectral Efficiency", "4G RAN", "90.0", ">=", "1.1"): 32,
        ("Voice Call Success Rate (VoLTE)", "4G RAN", "95.0", ">=", "97.0"): 34,
        ("Voice Call Drop Rate (VoLTE)", "4G RAN", "95.0", "<", "2.0"): 35,
        ("SRVCC Success Rate", "4G RAN", "95.0", ">=", "97.0"): 36,
        ("CQI", "5G RAN", "30.0", ">=", "10.0"): 42,
        ("Packet Latency", "5G RAN", "95.0", "<", "20.0"): 44,
    }

    def __init__(self):
        # Background fills (light green/red) - untuk kedua cells
        self.pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Text colors (dark green/red) - kontras dengan background
        self.pass_font = Font(color="228B22", bold=True)  # Forest Green
        self.fail_font = Font(color="DC143C", bold=True)  # Crimson
        
        self.center_align = Alignment(horizontal="center", vertical="center")

    def write_report(self, report: ClusterReport, output_path: Path) -> None:
        """
        Write report to Excel using template.
        Expects 'datatemplate.xlsx' in current directory.
        """
        template_path = Path("datatemplate.xlsx")
        if not template_path.exists():
            raise FileNotFoundError(f"Template file not found: {template_path}")

        # Copy template to output
        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy(template_path, output_path)

        # Load workbook
        wb = load_workbook(output_path)
        if "Template" not in wb.sheetnames:
            raise ValueError("Sheet 'Template' not found in template file")

        ws = wb["Template"]
        
        # Rename sheet from "Template" to "FAC-BAU"
        ws.title = "FAC-BAU"

        # Organize results by (kpi_name, tech, target_pct, op, thresh) -> {month: result}
        result_map = {}
        for r in report.kpi_results:
            key = self._make_key(r)
            if key not in result_map:
                result_map[key] = {}
            result_map[key][r.month] = r

        # Populate cells M-R for each KPI row
        for key, row_num in self.KPI_ROW_MAP.items():
            if key not in result_map:
                continue

            month_data = result_map[key]
            months_sorted = sorted(month_data.keys())

            # We expect up to 3 months
            for idx, month in enumerate(months_sorted[:3]):
                result = month_data[month]
                # Columns for month idx: M,N (month1), O,P (month2), Q,R (month3)
                val_col = 13 + idx * 2  # M=13, O=15, Q=17
                stat_col = val_col + 1

                # Write value
                val_cell = ws.cell(row=row_num, column=val_col)
                val_cell.value = f"{result.achievement_percentage:.2f}%"
                val_cell.alignment = self.center_align

                # Write status
                stat_cell = ws.cell(row=row_num, column=stat_col)
                status_text = "PASS" if result.passed else "FAIL"
                stat_cell.value = status_text
                stat_cell.alignment = self.center_align

                # Color coding - BACKGROUND + TEXT untuk SEMUA cells
                if result.passed:
                    # GOOD: Light green background + dark green text
                    val_cell.fill = self.pass_fill
                    val_cell.font = self.pass_font
                    stat_cell.fill = self.pass_fill
                    stat_cell.font = self.pass_font
                else:
                    # BAD: Light red background + dark red text
                    val_cell.fill = self.fail_fill
                    val_cell.font = self.fail_font
                    stat_cell.fill = self.fail_fill
                    stat_cell.font = self.fail_font

        # Save workbook
        wb.save(output_path)

    def _make_key(self, result: KPIResult) -> tuple:
        """
        Create a key tuple to match KPI_ROW_MAP.
        Format: (kpi_name, technology, target_pct_str, operator, threshold_str)
        """
        t = result.target
        # Convert to strings with 1 decimal
        target_pct_str = f"{t.target_percentage:.1f}"
        threshold_str = f"{t.threshold_value:.1f}" if isinstance(t.threshold_value, float) else str(t.threshold_value)
        
        return (t.name, t.technology, target_pct_str, t.operator, threshold_str)
