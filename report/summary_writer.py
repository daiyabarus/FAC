"""Summary report writer - FIXED GROUPING ISSUE

The problem: groupby was using column index instead of column name
Solution: Use column name directly
"""

import os
import traceback
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config.settings import LTEColumns, GSMColumns


class SummaryExcelWriter:
    """Write summary report per TOWER_ID and CELL_NAME"""

    def __init__(self, output_path, validation_results, kpi_data, period_info=None):
        self.output_path = output_path
        self.validation_results = validation_results
        self.kpi_data = kpi_data
        self.period_info = period_info
        
        # Formatting
        self.pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.pass_font = Font(color="006100", bold=True)
        self.fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.fail_font = Font(color="9C0006", bold=True)
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_align = Alignment(horizontal='center', vertical='center')

    def write_summary(self):
        """Create summary Excel with 3 sheets: 4G, 2G, NGI"""
        print("\n=== Writing Cell-Level Summary Report ===")
        
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # Create sheets
        self._create_4g_sheet(wb)
        self._create_2g_sheet(wb)
        self._create_ngi_sheet(wb)
        
        # Save
        self._save_workbook(wb)

    def _create_4g_sheet(self, wb):
        """Create 4G summary sheet per TOWER_ID & CELL_NAME"""
        ws = wb.create_sheet("4G Summary", 0)
        
        headers = [
            "TOWER_ID", "CELL_NAME", "BAND",
            "Session Setup Success Rate",
            "RACH Success Rate (>= 85%)",
            "RACH Success Rate (< 55%)",
            "Handover Success Rate",
            "E-RAB Drop Rate",
            "DL Throughput (>= 3 Mbps)",
            "DL Throughput (< 1 Mbps)",
            "UL Throughput (>= 1 Mbps)",
            "UL Throughput (< 0.256 Mbps)",
            "UL Packet Loss",
            "DL Packet Loss",
            "CQI",
            "MIMO Rank2 (>= 35%)",
            "MIMO Rank2 (< 20%)",
            "UL RSSI",
            "Latency (< 30ms)",
            "Latency (30-40ms)",
            "LTC Non Capacity",
            "Overlap Rate",
            "SE 850MHz 2T2R",
            "SE 900MHz 2T2R",
            "SE 2100MHz 2T2R",
            "SE 1800MHz 2T2R",
            "SE 1800MHz 4T4R/8T8R",
            "SE 2100MHz 4T4R/8T8R",
            "SE 2300MHz 32T32R",
            "VoLTE CSSR",
            "VoLTE Drop",
            "SRVCC SR",
        ]
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_align
        
        # Get LTE data
        lte_data = self.kpi_data.get("lte")
        if lte_data is None or len(lte_data) == 0:
            print("  ⚠ No LTE data for 4G summary")
            return
        
        # Get period data
        lte_period_data = self._get_period_data(lte_data)
        if len(lte_period_data) == 0:
            print("  ⚠ No LTE data in period range")
            return
        
        print(f"  Using {len(lte_period_data)} LTE records from 3 periods")
        
        # 🔧 DEBUG: Check required columns
        required_cols = ["TOWER_ID", "LTE_BAND"]
        missing = [col for col in required_cols if col not in lte_period_data.columns]
        if missing:
            print(f"  ⚠ Missing columns in LTE data: {missing}")
            print(f"  Available columns: {list(lte_period_data.columns)}")
            return
        
        # Get CELL_NAME column
        cell_name_col = lte_period_data.columns[LTEColumns.CELL_NAME]
        print(f"  Using CELL_NAME column: '{cell_name_col}'")
        
        # Collect cell-level results
        cell_results = self._collect_4g_cell_results(lte_period_data, cell_name_col)
        
        # Write data rows
        row_idx = 2
        for cell_data in cell_results:
            ws.cell(row=row_idx, column=1).value = cell_data.get("TOWER_ID", "")
            ws.cell(row=row_idx, column=2).value = cell_data.get("CELL_NAME", "")
            ws.cell(row=row_idx, column=3).value = cell_data.get("BAND", "")
            
            # Write PASS/FAIL for each KPI
            col_idx = 4
            for kpi_key in self._get_4g_kpi_keys():
                is_pass = cell_data.get(kpi_key, None)
                cell = ws.cell(row=row_idx, column=col_idx)
                
                if is_pass is None:
                    cell.value = "N/A"
                else:
                    cell.value = "PASS" if is_pass else "FAIL"
                    cell.fill = self.pass_fill if is_pass else self.fail_fill
                    cell.font = self.pass_font if is_pass else self.fail_font
                
                cell.border = self.border
                cell.alignment = self.center_align
                col_idx += 1
            
            row_idx += 1
        
        # Auto-width columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        
        print(f"✓ Written 4G summary: {len(cell_results)} cells")

    def _create_2g_sheet(self, wb):
        """Create 2G summary sheet per TOWER_ID & CELL_NAME"""
        ws = wb.create_sheet("2G Summary", 1)
        
        headers = [
            "TOWER_ID",
            "CELL_NAME",
            "Call Setup Success Rate",
            "SDCCH Success Rate",
            "Drop Rate",
        ]
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_align
        
        # Get GSM data
        gsm_data = self.kpi_data.get("gsm")
        if gsm_data is None or len(gsm_data) == 0:
            print("  ⚠ No GSM data for 2G summary")
            return
        
        gsm_period_data = self._get_period_data(gsm_data)
        if len(gsm_period_data) == 0:
            print("  ⚠ No GSM data in period range")
            return
        
        print(f"  Using {len(gsm_period_data)} GSM records from 3 periods")
        
        # Get BTS_NAME column
        bts_name_col = gsm_period_data.columns[GSMColumns.BTS_NAME]
        print(f"  Using BTS_NAME column: '{bts_name_col}'")
        
        # Collect cell-level results
        cell_results = self._collect_2g_cell_results(gsm_period_data, bts_name_col)
        
        # Write data rows
        row_idx = 2
        for cell_data in cell_results:
            ws.cell(row=row_idx, column=1).value = cell_data.get("TOWER_ID", "")
            ws.cell(row=row_idx, column=2).value = cell_data.get("CELL_NAME", "")
            
            col_idx = 3
            for kpi_key in ["cssr", "sdcch_sr", "drop_rate"]:
                is_pass = cell_data.get(kpi_key, None)
                cell = ws.cell(row=row_idx, column=col_idx)
                
                if is_pass is None:
                    cell.value = "N/A"
                else:
                    cell.value = "PASS" if is_pass else "FAIL"
                    cell.fill = self.pass_fill if is_pass else self.fail_fill
                    cell.font = self.pass_font if is_pass else self.fail_font
                
                cell.border = self.border
                cell.alignment = self.center_align
                col_idx += 1
            
            row_idx += 1
        
        # Auto-width
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        print(f"✓ Written 2G summary: {len(cell_results)} cells")

    def _create_ngi_sheet(self, wb):
        """Create NGI summary sheet per TOWER_ID & CELL_NAME"""
        ws = wb.create_sheet("NGI Summary", 2)
        
        headers = [
            "TOWER_ID",
            "CELL_NAME",
            "RSRP (Urban)",
            "RSRP (Suburban)",
            "RSRQ (Urban)",
            "RSRQ (Suburban)",
        ]
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_align
        
        # Get NGI data
        ngi_data = self.kpi_data.get("ngi")
        if ngi_data is None or len(ngi_data) == 0:
            print("  ⚠ No NGI data for NGI summary")
            return
        
        # Collect cell-level results
        cell_results = self._collect_ngi_cell_results(ngi_data)
        
        # Write data rows
        row_idx = 2
        for cell_data in cell_results:
            ws.cell(row=row_idx, column=1).value = cell_data.get("TOWER_ID", "")
            ws.cell(row=row_idx, column=2).value = cell_data.get("CELL_NAME", "")
            
            col_idx = 3
            for kpi_key in ["rsrp_urban", "rsrp_suburban", "rsrq_urban", "rsrq_suburban"]:
                is_pass = cell_data.get(kpi_key, None)
                cell = ws.cell(row=row_idx, column=col_idx)
                
                if is_pass is None:
                    cell.value = "N/A"
                else:
                    cell.value = "PASS" if is_pass else "FAIL"
                    cell.fill = self.pass_fill if is_pass else self.fail_fill
                    cell.font = self.pass_font if is_pass else self.fail_font
                
                cell.border = self.border
                cell.alignment = self.center_align
                col_idx += 1
            
            row_idx += 1
        
        # Auto-width
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        print(f"✓ Written NGI summary: {len(cell_results)} cells")

    # ================================================================== #
    # PERIOD FILTERING
    # ================================================================== #

    def _get_period_data(self, df):
        """Get data from the 3 periods only (Period 1, 2, 3)"""
        if "PERIOD" not in df.columns:
            print("  ⚠ Warning: No PERIOD column, using all data")
            return df
        
        valid_periods = ["Period 1", "Period 2", "Period 3"]
        filtered = df[df["PERIOD"].isin(valid_periods)].copy()
        
        orig_len = len(df)
        filtered_len = len(filtered)
        print(f"  Filtered from {orig_len} to {filtered_len} rows (90-day period)")
        
        return filtered

    # ================================================================== #
    # COLLECTORS - FIXED
    # ================================================================== #

    def _collect_4g_cell_results(self, lte_data, cell_name_col):
        """
        Collect PASS/FAIL per cell for 4G KPIs
        
        🔧 FIX: Pass cell_name_col explicitly to avoid column index issues
        """
        results = []
        
        # 🔧 DEBUG: Check if we can group
        try:
            # Group by TOWER_ID + CELL_NAME + BAND
            grouped = lte_data.groupby(["TOWER_ID", cell_name_col, "LTE_BAND"])
            print(f"  Grouped into {len(grouped)} unique cells")
        except Exception as e:
            print(f"  ✗ Error grouping LTE data: {e}")
            return results
        
        for (tower_id, cell_name, band), group in grouped:
            cell_result = {
                "TOWER_ID": tower_id,
                "CELL_NAME": cell_name,
                "BAND": band,
            }
            
            # Check each KPI
            cell_result["session_ssr"] = self._check_kpi(group, "SESSION_SSR", 99, ">=")
            cell_result["rach_sr_high"] = self._check_kpi(group, "RACH_SR", 85, ">=")
            cell_result["rach_sr_low"] = self._check_kpi(group, "RACH_SR", 55, ">=")
            cell_result["ho_sr"] = self._check_kpi(group, "HO_SR", 97, ">=")
            cell_result["erab_drop"] = self._check_kpi(group, "ERAB_DROP", 2, "<")
            cell_result["dl_thp_high"] = self._check_kpi(group, "DL_THP", 3, ">=")
            cell_result["dl_thp_low"] = self._check_kpi(group, "DL_THP", 1, ">=")
            cell_result["ul_thp_high"] = self._check_kpi(group, "UL_THP", 1, ">=")
            cell_result["ul_thp_low"] = self._check_kpi(group, "UL_THP", 0.256, ">=")
            cell_result["ul_ploss"] = self._check_kpi(group, "UL_PLOSS", 0.85, "<")
            cell_result["dl_ploss"] = self._check_kpi(group, "DL_PLOSS", 0.10, "<")
            cell_result["cqi"] = self._check_kpi(group, "CQI", 7, ">=")
            cell_result["mimo_rank2_high"] = self._check_kpi(group, "MIMO_RANK2", 35, ">=")
            cell_result["mimo_rank2_low"] = self._check_kpi(group, "MIMO_RANK2", 20, ">=")
            cell_result["ul_rssi"] = self._check_kpi(group, "UL_RSSI", -105, "<=")
            cell_result["latency_low"] = self._check_kpi(group, "LATENCY", 30, "<")
            cell_result["latency_medium"] = self._check_kpi(group, "LATENCY", 40, "<")
            cell_result["ltc_non_cap"] = self._check_kpi(group, "LTC_NON_CAP", 3, "<")
            cell_result["overlap_rate"] = self._check_kpi(group, "OVERLAP_RATE", 35, "<")
            
            # SE checks
            cell_result["se_850_2t2r"] = self._check_se(group, "2T2R", 850, 1.1)
            cell_result["se_900_2t2r"] = self._check_se(group, "2T2R", 900, 1.1)
            cell_result["se_2100_2t2r"] = self._check_se(group, "2T2R", 2100, 1.3)
            cell_result["se_1800_2t2r"] = self._check_se(group, "2T2R", 1800, 1.25)
            cell_result["se_1800_4t4r"] = self._check_se(group, ["4T4R", "8T8R"], 1800, 1.5)
            cell_result["se_2100_4t4r"] = self._check_se(group, ["4T4R", "8T8R"], [2100, 2300], 1.7)
            cell_result["se_2300_32t32r"] = self._check_se(group, "32T32R", 2300, 2.1)
            
            # VoLTE
            cell_result["volte_cssr"] = self._check_kpi(group, "VOLTE_CSSR", 97, ">=")
            cell_result["volte_drop"] = self._check_kpi(group, "VOLTE_DROP", 2, "<")
            cell_result["srvcc_sr"] = self._check_kpi(group, "SRVCC_SR", 97, ">=")
            
            results.append(cell_result)
        
        return results

    def _collect_2g_cell_results(self, gsm_data, bts_name_col):
        """
        Collect PASS/FAIL per cell for 2G KPIs
        
        🔧 FIX: Pass bts_name_col explicitly
        """
        results = []
        
        try:
            grouped = gsm_data.groupby(["TOWER_ID", bts_name_col])
            print(f"  Grouped into {len(grouped)} unique cells")
        except Exception as e:
            print(f"  ✗ Error grouping GSM data: {e}")
            return results
        
        for (tower_id, cell_name), group in grouped:
            cell_result = {
                "TOWER_ID": tower_id,
                "CELL_NAME": cell_name,
                "cssr": self._check_kpi(group, "CSSR", 98.5, ">="),
                "sdcch_sr": self._check_kpi(group, "SDCCH_SR", 98.5, ">="),
                "drop_rate": self._check_kpi(group, "DROP_RATE", 2, "<"),
            }
            results.append(cell_result)
        
        return results

    def _collect_ngi_cell_results(self, ngi_data):
        """Collect PASS/FAIL per cell for NGI"""
        results = []
        
        for _, row in ngi_data.iterrows():
            tower_id = row.get("TOWER_ID", "")
            cell_name = row.get("CELL_NAME", "")
            cat = str(row.get("CAT", "")).upper()
            rsrp = row.get("RSRP")
            rsrq = row.get("RSRQ")
            
            existing = next((r for r in results if r["TOWER_ID"] == tower_id and r["CELL_NAME"] == cell_name), None)
            
            if existing is None:
                existing = {
                    "TOWER_ID": tower_id,
                    "CELL_NAME": cell_name,
                    "rsrp_urban": None,
                    "rsrp_suburban": None,
                    "rsrq_urban": None,
                    "rsrq_suburban": None,
                }
                results.append(existing)
            
            if cat == "URBAN":
                existing["rsrp_urban"] = rsrp >= -105 if pd.notna(rsrp) else None
                existing["rsrq_urban"] = rsrq >= -12 if pd.notna(rsrq) else None
            elif cat == "SUBURBAN":
                existing["rsrp_suburban"] = rsrp >= -110 if pd.notna(rsrp) else None
                existing["rsrq_suburban"] = rsrq >= -14 if pd.notna(rsrq) else None
        
        return results

    def _check_kpi(self, group, column, threshold, operator):
        """Check if KPI passes threshold"""
        if column not in group.columns:
            return None
        
        values = group[column].dropna()
        if len(values) == 0:
            return None
        
        avg = values.mean()
        
        if operator == ">=":
            return avg >= threshold
        elif operator == ">":
            return avg > threshold
        elif operator == "<":
            return avg < threshold
        elif operator == "<=":
            return avg <= threshold
        else:
            return None

    def _check_se(self, group, tx_cond, band_cond, threshold):
        """Check spectral efficiency"""
        if "TX" not in group.columns or "LTE_BAND" not in group.columns or "SPECTRAL_EFF" not in group.columns:
            return None
        
        if isinstance(tx_cond, list):
            mask_tx = group["TX"].isin(tx_cond)
        else:
            mask_tx = group["TX"] == tx_cond
        
        if isinstance(band_cond, list):
            mask_band = group["LTE_BAND"].isin(band_cond)
        else:
            mask_band = group["LTE_BAND"] == band_cond
        
        filtered = group[mask_tx & mask_band]
        se_vals = filtered["SPECTRAL_EFF"].dropna()
        
        if len(se_vals) == 0:
            return None
        
        return se_vals.mean() >= threshold

    def _get_4g_kpi_keys(self):
        """Get ordered list of 4G KPI keys"""
        return [
            "session_ssr", "rach_sr_high", "rach_sr_low", "ho_sr", "erab_drop",
            "dl_thp_high", "dl_thp_low", "ul_thp_high", "ul_thp_low",
            "ul_ploss", "dl_ploss", "cqi", "mimo_rank2_high", "mimo_rank2_low",
            "ul_rssi", "latency_low", "latency_medium", "ltc_non_cap", "overlap_rate",
            "se_850_2t2r", "se_900_2t2r", "se_2100_2t2r", "se_1800_2t2r",
            "se_1800_4t4r", "se_2100_4t4r", "se_2300_32t32r",
            "volte_cssr", "volte_drop", "srvcc_sr",
        ]

    def _save_workbook(self, wb):
        """Save workbook"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"{self.output_path}/Summary_FAC_Report_{timestamp}.xlsx"
            
            if os.path.exists(output_file):
                os.remove(output_file)
            
            wb.save(output_file)
            wb.close()
            
            print(f"✓ Summary cell report saved: {output_file}")
            
            if os.path.exists(output_file):
                file_size = os.path.getsize(output_file)
                print(f"✓ File size: {file_size:,} bytes")
        except Exception as e:
            print(f"✗ Error saving summary report: {e}")
            traceback.print_exc()
            raise