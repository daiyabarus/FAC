"""Excel report writer - PERIOD based (Period 1–3)"""

import os
import base64
import traceback
from io import BytesIO
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment

from report.formatter import ExcelFormatter
from report.chart_generator import ChartGenerator
from utils.helpers import (
    format_date_range,
    get_three_month_range,
    format_period_date_range,
)
from config.settings import LTEColumns, GSMColumns


class ExcelReportWriter:
    """Write reports to Excel files"""

    def __init__(
        self,
        template_path,
        output_path,
        validation_results,
        kpi_data,
        transformed_data,
        period_info=None,
    ):
        self.template_path = template_path
        self.output_path = output_path
        self.validation_results = validation_results
        self.kpi_data = kpi_data          # raw/enriched LTE/GSM for RAW & charts
        self.transformed_data = transformed_data  # includes PERIOD, NGI, etc.
        self.formatter = ExcelFormatter()
        self.xlsmart_logo = None
        self.zte_logo = None
        self.period_info = period_info

    # ------------------------------------------------------------------ #
    # BASIC HELPERS
    # ------------------------------------------------------------------ #

    def _get_sorted_periods(self, cluster):
        """
        Get 3 periods in order (Period 1, Period 2, Period 3).
        Saat ini diasumsikan selalu 3 period tetap.
        """
        return ["Period 1", "Period 2", "Period 3"]

    def set_logos(self, xlsmart_logo, zte_logo):
        """Set logo base64 strings"""
        self.xlsmart_logo = xlsmart_logo
        self.zte_logo = zte_logo

    # ------------------------------------------------------------------ #
    # MAIN ENTRY
    # ------------------------------------------------------------------ #

    def write_report(self, cluster):
        """Write report for a specific cluster"""
        print(f"\n=== Writing report for cluster: {cluster} ===")
        wb = self._load_template()

        periods = self._get_sorted_periods(cluster)
        if not periods:
            print(f"⚠ No data for cluster {cluster}")
            return

        self._write_all_sheets(wb, cluster, periods)
        self._save_workbook(wb, cluster)

    # ------------------------------------------------------------------ #
    # TEMPLATE / SAVE
    # ------------------------------------------------------------------ #

    def _load_template(self):
        """Load Excel template or create new workbook"""
        try:
            wb = load_workbook(self.template_path)
            print(f"✓ Template loaded: {self.template_path}")
        except Exception as e:
            print(f"⚠ Could not load template: {e}")
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
        return wb

    def _save_workbook(self, wb, cluster):
        """Save workbook with proper naming and error handling"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"{self.output_path}/FAC_Report_{cluster}_{timestamp}.xlsx"

            if os.path.exists(output_file):
                os.remove(output_file)

            wb.save(output_file)
            wb.close()
            print(f"✓ Report saved: {output_file}")

            if os.path.exists(output_file):
                file_size = os.path.getsize(output_file)
                print(f"✓ File size: {file_size:,} bytes")
        except Exception as e:
            print(f"✗ Error saving report: {e}")
            traceback.print_exc()
            raise

    # ------------------------------------------------------------------ #
    # SHEET DISPATCH
    # ------------------------------------------------------------------ #

    def _write_all_sheets(self, wb, cluster, periods):
        """Write all sheets with error handling"""
        cluster_results = self.validation_results.get(cluster, {})

        sheets = [
            (
                "FAC",
                lambda: self._write_fac_sheet(
                    wb, cluster, periods, cluster_results
                ),
            ),
            (
                "Contributors",
                lambda: self._write_contributors_sheet(
                    wb, cluster, periods
                ),
            ),
            (
                "NGI Contributors",
                lambda: self._write_ngi_contributors_sheet(wb, cluster),
            ),
            ("RAW", lambda: self._write_raw_sheets(wb, cluster)),
            ("Charts", lambda: self._write_charts_sheet(wb, cluster)),
        ]

        for sheet_name, write_func in sheets:
            try:
                write_func()
                print(f"✓ {sheet_name} sheet written")
            except Exception as e:
                print(f"✗ Error writing {sheet_name} sheet: {e}")
                traceback.print_exc()

    # ------------------------------------------------------------------ #
    # FAC / TEMPLATE SHEET
    # ------------------------------------------------------------------ #

    def _write_fac_sheet(self, wb, cluster, periods, cluster_results):
        """Write FAC/Template sheet"""
        if "Template" in wb.sheetnames:
            ws = wb["Template"]
        elif "FAC" in wb.sheetnames:
            ws = wb["FAC"]
        else:
            ws = wb.create_sheet("FAC")

        ws.title = "FAC"
        self._add_logos(ws)

        lte_data = self.transformed_data["lte"]
        gsm_data = self.transformed_data["gsm"]
        cluster_lte = lte_data[lte_data["CLUSTER"] == cluster]
        cluster_gsm = gsm_data[gsm_data["CLUSTER"] == cluster]

        dates = pd.to_datetime(
            cluster_lte.iloc[:, LTEColumns.BEGIN_TIME]
        ).unique()
        date_range = get_three_month_range(dates)

        ws["A6"] = f"FAC KPI Achievement {cluster} - {date_range}"
        ws["A6"].font = self.formatter.header_font

        self._write_cluster_info(ws, cluster, cluster_lte, cluster_gsm, dates)

        periods = periods[:3]
        self._write_month_headers(ws, periods, cluster_lte)
        self._write_kpi_results(ws, periods, cluster_results)
        self._write_overall_results(ws, periods, cluster_results)

    def _write_cluster_info(self, ws, cluster, cluster_lte, cluster_gsm, dates):
        """Write cluster information to A14"""
        unique_towers = cluster_lte["TOWER_ID"].dropna().nunique()
        unique_4g_cells = (
            cluster_lte.iloc[:, LTEColumns.CELL_NAME].dropna().nunique()
        )
        unique_2g_cells = (
            cluster_gsm.iloc[:, GSMColumns.BTS_NAME].dropna().nunique()
        )

        if len(dates) > 0:
            last_date = pd.to_datetime(max(dates))
            last_pac_date = last_date.strftime("%d %B %Y")
        else:
            last_pac_date = "N/A"

        info_text = (
            f"City name: {cluster}\n"
            f"Site Number: {unique_towers}\n"
            f"Cell Number: {unique_4g_cells} 4G & {unique_2g_cells} 2G\n"
            f"LAST CLUSTER PAC DATE: {last_pac_date}"
        )

        ws["A14"] = info_text
        ws["A14"].alignment = Alignment(
            wrap_text=True, vertical="top", horizontal="left"
        )

    # def _write_month_headers(self, ws, periods, cluster_lte):
    #     """Write period headers to row 12 and date ranges to row 13"""
    #     for idx, period in enumerate(periods):
    #         col = 13 + (idx * 2)

    #         if self.period_info:
    #             try:
    #                 period_num = int(period.split()[1])
    #                 period_data = self.period_info[f"period_{period_num}"]
    #                 period_label = period_data.get("label", period)
    #                 date_range = format_period_date_range(period_data)
                    
    #                 # 🐛 DEBUG PRINT
    #                 print(f"  DEBUG Period {period_num}:")
    #                 print(f"    period_data: {period_data}")
    #                 print(f"    label: {period_label}")
    #                 print(f"    date_range: {date_range}")
                    
    #             except Exception as e:
    #                 print(f"  ⚠ Error formatting period {period}: {e}")
    #                 traceback.print_exc()
    #                 period_label = period
    #                 date_range = "N/A"
    #         else:
    #             print(f"  ⚠ No period_info available!")
    #             period_label = period
    #             date_range = "N/A"

    #         ws.cell(row=12, column=col).value = period_label
    #         ws.cell(row=12, column=col).font = self.formatter.header_font
    #         ws.cell(row=14, column=col).value = date_range

    #         print(f"  Period {idx + 1} written: {period_label} (col {col})")
    #         print(f"    Date range: {date_range}")

    def _write_month_headers(self, ws, periods, cluster_lte):
        """Write period date ranges to row 13 only (row 12 from template)"""
        for idx, period in enumerate(periods):
            col = 13 + (idx * 2)
            
            if self.period_info:
                try:
                    period_num = int(period.split()[1])
                    period_data = self.period_info[f"period_{period_num}"]
                    date_range = format_period_date_range(period_data)
                    ws.cell(row=13, column=col).value = date_range
                    
                    print(f"  Period {idx + 1} date range: {date_range} (column {col})")
                    
                except Exception as e:
                    print(f"  ⚠ Error formatting period {period}: {e}")
                    traceback.print_exc()
                    ws.cell(row=13, column=col).value = "N/A"
            else:
                print(f"  ⚠ No period_info available!")
                ws.cell(row=13, column=col).value = "N/A"

    def _write_overall_results(self, ws, periods, cluster_results):
        """Write overall pass/fail results to row 63"""
        for idx, period in enumerate(periods):
            col = 13 + (idx * 2)
            period_results = cluster_results.get(period, {})
            overall_pass = period_results.get("overall_pass", False)
            result_cell = ws.cell(row=63, column=col)
            self.formatter.format_pass_fail(result_cell, overall_pass)

    def _write_kpi_results(self, ws, periods, cluster_results):
        """Write KPI results including NGI"""
        kpi_mapping = self._get_kpi_mapping()

        for idx, period in enumerate(periods):
            value_col = 13 + idx * 2
            result_col = value_col + 1

            period_results = cluster_results.get(period, {})
            # GSM
            self._write_tech_kpi_results(
                ws,
                period_results.get("gsm", {}),
                kpi_mapping,
                "gsm",
                value_col,
                result_col,
            )
            # LTE
            self._write_tech_kpi_results(
                ws,
                period_results.get("lte", {}),
                kpi_mapping,
                "lte",
                value_col,
                result_col,
            )

        # NGI (single section, not per period)
        ngi_results = cluster_results.get("NGI")
        if ngi_results:
            print(f"  Writing NGI results: {list(ngi_results.keys())}")
            value_col = 13
            status_col = 16

            for key, kpi in ngi_results.items():
                row = kpi.get("row")
                if row is None:
                    continue

                value_cell = ws.cell(row=row, column=value_col)
                self.formatter.format_value(value_cell, kpi["value"], "0.00")
                status_cell = ws.cell(row=row, column=status_col)
                self.formatter.format_pass_fail(status_cell, kpi["pass"])

                print(
                    f"    ✓ {key} written to row {row}: "
                    f"{kpi['value']:.2f}% - "
                    f"{'PASS' if kpi['pass'] else 'FAIL'}"
                )

    def _get_kpi_mapping(self):
        """Get KPI to row mapping"""
        return {
            # GSM KPIs
            ("gsm", "cssr"): 14,
            ("gsm", "sdcch_sr"): 15,
            ("gsm", "drop_rate"): 16,
            # LTE KPIs
            ("lte", "session_ssr"): 17,
            ("lte", "rach_sr_high"): 18,
            ("lte", "rach_sr_low"): 19,
            ("lte", "ho_sr"): 20,
            ("lte", "erab_drop"): 21,
            ("lte", "dl_thp_high"): 22,
            ("lte", "dl_thp_low"): 23,
            ("lte", "ul_thp_high"): 24,
            ("lte", "ul_thp_low"): 25,
            ("lte", "ul_ploss"): 26,
            ("lte", "dl_ploss"): 27,
            ("lte", "cqi"): 28,
            ("lte", "mimo_rank2_high"): 29,
            ("lte", "mimo_rank2_low"): 30,
            ("lte", "ul_rssi"): 31,
            ("lte", "latency_low"): 32,
            ("lte", "latency_medium"): 33,
            ("lte", "ltc_non_cap"): 34,
            ("lte", "overlap_rate"): 35,
            # Spectral Efficiency
            ("lte", "se_850_2t2r"): 36,
            ("lte", "se_900_2t2r"): 37,
            ("lte", "se_2100_2t2r"): 38,
            ("lte", "se_1800_2t2r"): 39,
            ("lte", "se_1800_4t4r"): 40,
            ("lte", "se_2100_4t4r"): 41,
            ("lte", "se_2300_32t32r"): 43,
            # VoLTE
            ("lte", "volte_cssr"): 45,
            ("lte", "volte_drop"): 46,
            ("lte", "srvcc_sr"): 47,
            # NGI (rows in FAC)
            ("ngi", "ngi_rsrp_urban"): 58,
            ("ngi", "ngi_rsrp_suburban"): 59,
            ("ngi", "ngi_rsrq_urban"): 60,
            ("ngi", "ngi_rsrq_suburban"): 61,
        }

    def _write_tech_kpi_results(
        self, ws, tech_results, kpi_mapping, tech, value_col, result_col
    ):
        """Write KPI results for a specific technology"""
        for kpi_key, kpi_result in tech_results.items():
            if kpi_key.startswith("se_"):
                row = kpi_result.get("row", None)
                if row is None:
                    print(f"  ⚠ Skipping {kpi_key} - no row mapping")
                    continue
            else:
                row_key = (tech, kpi_key)
                if row_key not in kpi_mapping:
                    continue
                row = kpi_mapping[row_key]

            value_cell = ws.cell(row=row, column=value_col)
            self.formatter.format_value(
                value_cell, kpi_result["value"], "0.00"
            )
            result_cell = ws.cell(row=row, column=result_col)
            self.formatter.format_pass_fail(result_cell, kpi_result["pass"])

            if kpi_key.startswith("se_"):
                print(
                    f"  ✓ Written {kpi_key} to row {row}: "
                    f"value={kpi_result['value']:.2f}%, "
                    f"pass={kpi_result['pass']}"
                )

    # ------------------------------------------------------------------ #
    # CONTRIBUTORS (PERIOD-BASED)
    # ------------------------------------------------------------------ #

    def _collect_gsm_contributors(self, gsm_cluster, period):
        """Collect GSM failed cells for a given period"""
        contributors = []
        gsm_period = gsm_cluster[gsm_cluster["PERIOD"] == period]

        gsm_checks = [
            ("CSSR", 98.5, "<", "Accessibility", "Call Setup Success Rate"),
            ("SDCCH_SR", 98.5, "<", "Accessibility", "SDCCH Success rate"),
            ("DROP_RATE", 2, ">=", "Retainability", "Perceive Drop Rate"),
        ]

        for kpi_col, baseline, operator, clause, name in gsm_checks:
            if kpi_col not in gsm_period.columns:
                continue

            if operator == "<":
                fails = gsm_period[gsm_period[kpi_col] < baseline]
            else:
                fails = gsm_period[gsm_period[kpi_col] >= baseline]

            for _, row in fails.iterrows():
                contributors.append(
                    {
                        "Month": period,
                        "Clause Type": clause,
                        "Name": name,
                        "Reference": "2G RAN",
                        "TOWER_ID": row.get("TOWER_ID", ""),
                        "CELL_NAME": row.iloc[GSMColumns.BTS_NAME],
                        "Value": row[kpi_col],
                        "Baseline": baseline,
                        "Status": "FAIL",
                    }
                )

        return contributors

    def _collect_lte_contributors(self, lte_cluster, period):
        """Collect LTE failed cells for a given period"""
        contributors = []
        lte_period = lte_cluster[lte_cluster["PERIOD"] == period]

        lte_checks = [
            ("SESSION_SSR", 99, "<", "Accessibility", "Session Setup Success Rate"),
            ("RACH_SR", 85, "<", "Accessibility", "RACH Success Rate (< 85%)"),
            ("RACH_SR", 55, "<", "Accessibility", "RACH Success Rate (< 55%)"),
            ("HO_SR", 97, "<", "Mobility",
             "Handover Success Rate Inter and Intra-Frequency"),
            ("ERAB_DROP", 2, ">=", "Retainability", "E-RAB Drop Rate"),
            ("DL_THP", 3, "<", "Integrity",
             "Downlink User Throughput (< 3 Mbps)"),
            ("DL_THP", 1, "<", "Integrity",
             "Downlink User Throughput (< 1 Mbps)"),
            ("UL_THP", 1, "<", "Integrity",
             "Uplink User Throughput (< 1 Mbps)"),
            ("UL_THP", 0.256, "<", "Integrity",
             "Uplink User Throughput (< 0.256 Mbps)"),
            ("UL_PLOSS", 0.85, ">=", "Integrity", "UL Packet Loss (PDCP)"),
            ("DL_PLOSS", 0.10, ">=", "Integrity", "DL Packet Loss (PDCP)"),
            ("CQI", 7, "<", "Integrity", "CQI"),
            ("MIMO_RANK2", 35, "<", "Integrity",
             "MIMO Transmission Rank2 Rate (< 35%)"),
            ("MIMO_RANK2", 20, "<", "Integrity",
             "MIMO Transmission Rank2 Rate (< 20%)"),
            ("LATENCY", 30, ">=", "Integrity", "Packet Latency (>= 30 ms)"),
            ("LATENCY", 40, ">=", "Integrity", "Packet Latency (>= 40 ms)"),
            ("LTC_NON_CAP", 3, "<", "Utilization", "LTC Non Capacity (<= 5%)"),
            ("VOLTE_CSSR", 97, "<", "VoLTE", "VoLTE Call Success Rate"),
            ("VOLTE_DROP", 2, ">=", "VoLTE", "VoLTE Call Drop Rate"),
            ("SRVCC_SR", 97, "<", "VoLTE", "SRVCC Success Rate"),
        ]

        for kpi_col, baseline, operator, clause, name in lte_checks:
            if kpi_col not in lte_period.columns:
                continue

            if operator == "<":
                fails = lte_period[lte_period[kpi_col] < baseline]
            else:
                fails = lte_period[lte_period[kpi_col] >= baseline]

            for _, row in fails.iterrows():
                contributors.append(
                    {
                        "Month": period,
                        "Clause Type": clause,
                        "Name": name,
                        "Reference": "4G RAN",
                        "TOWER_ID": row.get("TOWER_ID", ""),
                        "CELL_NAME": row.iloc[LTEColumns.CELL_NAME],
                        "Value": row[kpi_col],
                        "Baseline": baseline,
                        "Status": "FAIL",
                    }
                )

        # UL_RSSI
        if "UL_RSSI" in lte_period.columns:
            rssi_fails = lte_period[lte_period["UL_RSSI"] > -105]
            for _, row in rssi_fails.iterrows():
                contributors.append(
                    {
                        "Month": period,
                        "Clause Type": "Integrity",
                        "Name": "UL RSSI (> -105 dBm)",
                        "Reference": "4G RAN",
                        "TOWER_ID": row.get("TOWER_ID", ""),
                        "CELL_NAME": row.iloc[LTEColumns.CELL_NAME],
                        "Value": row["UL_RSSI"],
                        "Baseline": -105,
                        "Status": "FAIL",
                    }
                )

        # OVERLAP_RATE
        if "OVERLAP_RATE" in lte_period.columns:
            ov_fails = lte_period[lte_period["OVERLAP_RATE"] >= 35]
            for _, row in ov_fails.iterrows():
                contributors.append(
                    {
                        "Month": period,
                        "Clause Type": "Coverage",
                        "Name": "Coverage Overlapping Ratio (< 35%)",
                        "Reference": "4G RAN",
                        "TOWER_ID": row.get("TOWER_ID", ""),
                        "CELL_NAME": row.iloc[LTEColumns.CELL_NAME],
                        "Value": row["OVERLAP_RATE"],
                        "Baseline": 35,
                        "Status": "FAIL",
                    }
                )

        # Spectral Efficiency
        if "TX" in lte_period.columns and "LTE_BAND" in lte_period.columns and "SPECTRAL_EFF" in lte_period.columns:
            se_checks = [
                ("se_850_2t2r", "2T2R", 850, 1.1,
                 "Spectral Efficiency 850MHz 2T2R"),
                ("se_900_2t2r", "2T2R", 900, 1.1,
                 "Spectral Efficiency 900MHz 2T2R"),
                ("se_2100_2t2r", "2T2R", 2100, 1.3,
                 "Spectral Efficiency 2100MHz 2T2R"),
                ("se_1800_2t2r", "2T2R", 1800, 1.25,
                 "Spectral Efficiency 1800MHz 2T2R"),
                (
                    "se_1800_4t4r",
                    ["4T4R", "8T8R"],
                    1800,
                    1.5,
                    "Spectral Efficiency 1800MHz 4T4R/8T8R",
                ),
                (
                    "se_2100_4t4r",
                    ["4T4R", "8T8R"],
                    [2100, 2300],
                    1.7,
                    "Spectral Efficiency 2100/2300MHz 4T4R/8T8R",
                ),
                (
                    "se_2300_32t32r",
                    "32T32R",
                    2300,
                    2.1,
                    "Spectral Efficiency 2300MHz 32T32R",
                ),
            ]

            for key, tx_cond, band_cond, baseline, name in se_checks:
                if isinstance(tx_cond, list):
                    mask_tx = lte_period["TX"].isin(tx_cond)
                else:
                    mask_tx = lte_period["TX"] == tx_cond

                if isinstance(band_cond, list):
                    mask_band = lte_period["LTE_BAND"].isin(band_cond)
                else:
                    mask_band = lte_period["LTE_BAND"] == band_cond

                filtered = lte_period[mask_tx & mask_band]
                se_vals = filtered["SPECTRAL_EFF"].dropna()

                if len(se_vals) == 0:
                    continue

                fails = filtered[filtered["SPECTRAL_EFF"] < baseline]

                for _, row in fails.iterrows():
                    contributors.append(
                        {
                            "Month": period,
                            "Clause Type": "Spectral Efficiency",
                            "Name": name,
                            "Reference": "4G RAN",
                            "TOWER_ID": row.get("TOWER_ID", ""),
                            "CELL_NAME": row.iloc[LTEColumns.CELL_NAME],
                            "Value": row["SPECTRAL_EFF"],
                            "Baseline": baseline,
                            "Status": "FAIL",
                        }
                    )

        return contributors

    def _remove_duplicate_contributors(self, contributors):
        """Remove duplicate contributors"""
        if not contributors:
            return []

        df = pd.DataFrame(contributors)
        df = df.drop_duplicates(
            subset=[
                "Month",
                "Clause Type",
                "Name",
                "Reference",
                "TOWER_ID",
                "CELL_NAME",
                "Baseline",
            ],
            keep="first",
        )

        return df.to_dict("records")

    def _write_contributor_rows(self, ws, contributors):
        """Write contributor rows to worksheet"""
        for row_idx, contrib in enumerate(contributors, 2):
            ws.cell(row=row_idx, column=1).value = contrib["Month"]
            ws.cell(row=row_idx, column=2).value = contrib["Clause Type"]
            ws.cell(row=row_idx, column=3).value = contrib["Name"]
            ws.cell(row=row_idx, column=4).value = contrib["Reference"]
            ws.cell(row=row_idx, column=5).value = contrib["TOWER_ID"]
            ws.cell(row=row_idx, column=6).value = contrib["CELL_NAME"]

            value_cell = ws.cell(row=row_idx, column=7)
            self.formatter.format_value(value_cell, contrib["Value"], "0.00")

            ws.cell(row=row_idx, column=8).value = contrib["Baseline"]

            status_cell = ws.cell(row=row_idx, column=9)
            self.formatter.format_pass_fail(status_cell, False)

    def _write_contributors_sheet(self, wb, cluster, periods):
        """Write Contributors sheet with cells that failed (period-based)"""
        if "Contributors" in wb.sheetnames:
            wb.remove(wb["Contributors"])
        ws = wb.create_sheet("Contributors")

        headers = [
            "Month",
            "Clause Type",
            "Name",
            "Reference",
            "TOWER_ID",
            "CELL_NAME",
            "Value",
            "Baseline",
            "Status",
        ]

        for col_idx, header in enumerate(headers, 1):
            self.formatter.format_header_small(
                ws.cell(row=1, column=col_idx), header
            )

        contributors = self._collect_contributors(cluster, periods)
        contributors = self._remove_duplicate_contributors(contributors)
        self._write_contributor_rows(ws, contributors)

        print(
            f"✓ Written {len(contributors)} contributors"
        )

    def _collect_contributors(self, cluster, periods):
        """Collect all failed cells as contributors (GSM/LTE only, NO NGI)"""
        contributors = []
        lte_data = self.kpi_data["lte"]
        gsm_data = self.kpi_data["gsm"]
        lte_cluster = lte_data[lte_data["CLUSTER"] == cluster]
        gsm_cluster = gsm_data[gsm_data["CLUSTER"] == cluster]

        for period in periods:
            contributors.extend(
                self._collect_gsm_contributors(gsm_cluster, period)
            )
            contributors.extend(
                self._collect_lte_contributors(lte_cluster, period)
            )

        contributors = self._remove_duplicate_contributors(contributors)
        return contributors

    def collect_ngi_contributors(self, cluster):
        """
        Cari cell NGI yang FAIL terhadap threshold RSRP/RSRQ.
        Returns list dengan kolom CAT untuk sheet NGI Contributors.
        
        🔧 ADDED: Remove duplicates based on TOWER_ID + CELL_NAME + CAT + Metric
        """
        contributors = []
        ngi = self.transformed_data.get("ngi")
        if ngi is None or len(ngi) == 0:
            return contributors

        df = ngi[ngi["CLUSTER"] == cluster].copy()
        if len(df) == 0:
            return contributors

        df["CAT"] = df["CAT"].astype(str).str.upper()

        # RSRP URBAN (FAIL: < -105)
        mask = (df["CAT"] == "URBAN") & (df["RSRP"] < -105)
        for _, row in df[mask].iterrows():
            contributors.append(
                {
                    "CAT": row["CAT"],
                    "Clause Type": "Coverage Quality",
                    "Name": "RSRP (Urban)",
                    "TOWER_ID": row.get("TOWER_ID", ""),
                    "CELL_NAME": row.get("CELL_NAME", ""),
                    "Value": row["RSRP"],
                    "Baseline": -105,
                    "Status": "FAIL",
                }
            )

        # RSRP SUBURBAN (FAIL: < -110)
        mask = (df["CAT"] == "SUBURBAN") & (df["RSRP"] < -110)
        for _, row in df[mask].iterrows():
            contributors.append(
                {
                    "CAT": row["CAT"],
                    "Clause Type": "Coverage Quality",
                    "Name": "RSRP (Suburban)",
                    "TOWER_ID": row.get("TOWER_ID", ""),
                    "CELL_NAME": row.get("CELL_NAME", ""),
                    "Value": row["RSRP"],
                    "Baseline": -110,
                    "Status": "FAIL",
                }
            )

        # RSRQ URBAN (FAIL: < -12)
        mask = (df["CAT"] == "URBAN") & (df["RSRQ"] < -12)
        for _, row in df[mask].iterrows():
            contributors.append(
                {
                    "CAT": row["CAT"],
                    "Clause Type": "Coverage Quality",
                    "Name": "RSRQ (Urban)",
                    "TOWER_ID": row.get("TOWER_ID", ""),
                    "CELL_NAME": row.get("CELL_NAME", ""),
                    "Value": row["RSRQ"],
                    "Baseline": -12,
                    "Status": "FAIL",
                }
            )

        # RSRQ SUBURBAN (FAIL: < -14)
        mask = (df["CAT"] == "SUBURBAN") & (df["RSRQ"] < -14)
        for _, row in df[mask].iterrows():
            contributors.append(
                {
                    "CAT": row["CAT"],
                    "Clause Type": "Coverage Quality",
                    "Name": "RSRQ (Suburban)",
                    "TOWER_ID": row.get("TOWER_ID", ""),
                    "CELL_NAME": row.get("CELL_NAME", ""),
                    "Value": row["RSRQ"],
                    "Baseline": -14,
                    "Status": "FAIL",
                }
            )

        # 🔧 REMOVE DUPLICATES
        contributors = self._remove_duplicate_ngi_contributors(contributors)
        
        return contributors


    def _remove_duplicate_ngi_contributors(self, contributors):
        """
        Remove duplicate NGI contributors based on unique combination of:
        - TOWER_ID
        - CELL_NAME
        - CAT
        - Name (metric name)
        - Baseline
        
        Keep the first occurrence (worst value if multiple exist)
        """
        if not contributors:
            return []
        
        import pandas as pd
        
        df = pd.DataFrame(contributors)
        
        # Remove duplicates, keeping first occurrence
        df = df.drop_duplicates(
            subset=["TOWER_ID", "CELL_NAME", "CAT", "Name", "Baseline"],
            keep="first"
        )
        
        print(f"  Removed {len(contributors) - len(df)} duplicate NGI contributors")
        
        return df.to_dict("records")


    def _write_ngi_contributors_sheet(self, wb, cluster):
        """Write NGI Contributors sheet (separate from monthly contributors)"""
        ngi_contribs = self.collect_ngi_contributors(cluster)

        if not ngi_contribs:
            print("  No NGI contributors (all cells passed)")
            return

        if "NGI Contributors" in wb.sheetnames:
            wb.remove(wb["NGI Contributors"])
        ws = wb.create_sheet("NGI Contributors")

        headers = [
            "CAT",
            "Clause Type",
            "Name",
            "TOWER_ID",
            "CELL_NAME",
            "Value",
            "Baseline",
            "Status",
        ]
        for col_idx, h in enumerate(headers, 1):
            self.formatter.format_header_small(
                ws.cell(row=1, column=col_idx), h
            )

        row_idx = 2
        for c in ngi_contribs:
            ws.cell(row=row_idx, column=1, value=c.get("CAT", ""))
            ws.cell(row=row_idx, column=2, value=c["Clause Type"])
            ws.cell(row=row_idx, column=3, value=c["Name"])
            ws.cell(row=row_idx, column=4, value=c["TOWER_ID"])
            ws.cell(row=row_idx, column=5, value=c["CELL_NAME"])

            vcell = ws.cell(row=row_idx, column=6)
            self.formatter.format_value(vcell, c["Value"], "0.00")

            ws.cell(row=row_idx, column=7, value=c["Baseline"])

            scell = ws.cell(row=row_idx, column=8)
            self.formatter.format_pass_fail(scell, False)
            row_idx += 1

        print(
            f"✓ Written {len(ngi_contribs)} NGI contributors"
        )

    def _write_raw_sheets(self, wb, cluster):
        """Write RAW 2G and RAW 4G sheets"""
        self._write_raw_2g_sheet(wb, cluster)
        self._write_raw_4g_sheet(wb, cluster)

    def _write_raw_2g_sheet(self, wb, cluster):
        """Write RAW 2G sheet - FILTERED TO PERIOD RANGE"""
        gsm_data = self.kpi_data["gsm"]
        gsm_cluster = gsm_data[gsm_data["CLUSTER"] == cluster].copy()

        if len(gsm_cluster) == 0:
            return

        if "PERIOD" in gsm_cluster.columns and self.period_info is not None:
            valid_periods = ["Period 1", "Period 2", "Period 3"]
            gsm_cluster = gsm_cluster[gsm_cluster["PERIOD"].isin(valid_periods)].copy()
            print(f"  Filtered RAW 2G to period range: {len(gsm_cluster)} records")

        if len(gsm_cluster) == 0:
            print("  ⚠ No 2G data in period range")
            return

        from config.settings import GSMColumns
        
        gsm_raw = gsm_cluster[
            [
                gsm_cluster.columns[GSMColumns.BEGIN_TIME],
                "TOWER_ID",
                gsm_cluster.columns[GSMColumns.BTS_NAME],
                "CSSR",
                "SDCCH_SR",
                "DROP_RATE",
            ]
        ].copy()

        gsm_raw.columns = [
            "BEGIN_TIME",
            "TOWER_ID",
            "BTS_NAME",
            "CSSR",
            "SDCCH_SR",
            "DROP_RATE",
        ]

        gsm_raw = gsm_raw.drop_duplicates()
        gsm_raw["BEGIN_TIME"] = pd.to_datetime(
            gsm_raw["BEGIN_TIME"]
        ).dt.strftime("%-m/%-d/%Y")

        if "RAW 2G" in wb.sheetnames:
            wb.remove(wb["RAW 2G"])
        ws = wb.create_sheet("RAW 2G")

        self._write_raw_data(ws, gsm_raw)
        print(f"✓ Written RAW 2G: {len(gsm_raw)} records")

    def _write_raw_4g_sheet(self, wb, cluster):
        """Write RAW 4G sheet - FILTERED TO PERIOD RANGE"""
        lte_data = self.kpi_data["lte"]
        lte_cluster = lte_data[lte_data["CLUSTER"] == cluster].copy()

        if len(lte_cluster) == 0:
            return

        if "PERIOD" in lte_cluster.columns and self.period_info is not None:
            valid_periods = ["Period 1", "Period 2", "Period 3"]
            lte_cluster = lte_cluster[lte_cluster["PERIOD"].isin(valid_periods)].copy()
            print(f"  Filtered RAW 4G to period range: {len(lte_cluster)} records")

        if len(lte_cluster) == 0:
            print("  ⚠ No 4G data in period range")    
            return

        from config.settings import LTEColumns
        
        lte_raw = lte_cluster[
            [
                lte_cluster.columns[LTEColumns.BEGIN_TIME],
                "TOWER_ID",
                lte_cluster.columns[LTEColumns.CELL_NAME],
                "LTE_BAND",
                "SESSION_SSR",
                "RACH_SR",
                "HO_SR",
                "ERAB_DROP",
                "DL_THP",
                "UL_THP",
                "UL_PLOSS",
                "DL_PLOSS",
                "CQI",
                "MIMO_RANK2",
                "UL_RSSI",
                "LATENCY",
                "LTC_NON_CAP",
                "OVERLAP_RATE",
                "SPECTRAL_EFF",
                "VOLTE_CSSR",
                "VOLTE_DROP",
                "SRVCC_SR",
            ]
        ].copy()

        lte_raw.columns = [
            "BEGIN_TIME",
            "TOWER_ID",
            "CELL_NAME",
            "BAND",
            "SESSION_SSR",
            "RACH_SR",
            "HO_SR",
            "ERAB_DROP",
            "DL_THP",
            "UL_THP",
            "UL_PLOSS",
            "DL_PLOSS",
            "CQI",
            "MIMO_RANK2",
            "UL_RSSI",
            "LATENCY",
            "LTC_NON_CAP",
            "OVERLAP_RATE",
            "SPECTRAL_EFF",
            "VOLTE_CSSR",
            "VOLTE_DROP",
            "SRVCC_SR",
        ]

        lte_raw = lte_raw.drop_duplicates()
        lte_raw["BEGIN_TIME"] = pd.to_datetime(
            lte_raw["BEGIN_TIME"]
        ).dt.strftime("%-m/%-d/%Y")

        if "RAW 4G" in wb.sheetnames:
            wb.remove(wb["RAW 4G"])
        ws = wb.create_sheet("RAW 4G")

        self._write_raw_data(ws, lte_raw)
        print(f"✓ Written RAW 4G: {len(lte_raw)} records")

    def _write_raw_data(self, ws, df):
        """Write raw data to worksheet"""
        for c_idx, col in enumerate(df.columns, 1):
            self.formatter.format_header_small(
                ws.cell(row=1, column=c_idx), col
            )

        for r_idx, row in enumerate(df.itertuples(index=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value

                if c_idx > 3 and isinstance(value, (int, float)):
                    cell.number_format = "0.00"
    # ------------------------------------------------------------------ #
    # CHARTS SHEET
    # ------------------------------------------------------------------ #

    def _write_charts_sheet(self, wb, cluster):
        """Write Charts sheet with KPI trend charts"""
        print(f"\n=== Generating charts for {cluster} ===")

        chart_gen = ChartGenerator(
            self.kpi_data,
            self.transformed_data,
            cluster,
            self.period_info,
        )
        charts = chart_gen.generate_all_charts()

        if not charts:
            print("⚠ No charts generated")
            return

        if "Charts" in wb.sheetnames:
            wb.remove(wb["Charts"])
        ws = wb.create_sheet("Charts")

        ws["A1"] = f"KPI Trend Charts - {cluster}"
        ws["A1"].font = self.formatter.header_font_small

        current_row = 3
        current_col = 1
        charts_per_row = 2
        chart_height = 20
        chart_width = 12
        chart_num = 0

        for chart_key, chart_base64 in charts.items():
            try:
                img_data = base64.b64decode(chart_base64)
                img = XLImage(BytesIO(img_data))
                img.width = 700
                img.height = 350
                col_offset = (current_col - 1) * chart_width
                col_letter = self._get_column_letter(col_offset + 1)
                cell_position = f"{col_letter}{current_row}"
                ws.add_image(img, cell_position)
                chart_num += 1
                current_col += 1
                if current_col > charts_per_row:
                    current_col = 1
                    current_row += chart_height

                print(
                    f"✓ Added chart {chart_num}: {chart_key} at {cell_position}"
                )

            except Exception as e:
                print(f"⚠ Could not add chart {chart_key}: {e}")

        print(f"✓ Written {chart_num} charts to Charts sheet")

    def _get_column_letter(self, col_num):
        """Convert column number to Excel column letter"""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(col_num % 26 + 65) + result
            col_num //= 26
        return result

    # ------------------------------------------------------------------ #
    # LOGOS
    # ------------------------------------------------------------------ #

    def _add_logos(self, ws):
        """Add logos to worksheet"""
        try:
            if self.xlsmart_logo:
                img_data = base64.b64decode(self.xlsmart_logo)
                img = XLImage(BytesIO(img_data))
                img.width, img.height = 200, 70
                ws.add_image(img, "A1")

            if self.zte_logo:
                img_data = base64.b64decode(self.zte_logo)
                img = XLImage(BytesIO(img_data))
                img.width, img.height = 100, 50
                ws.add_image(img, "R1")
        except Exception as e:
            print(f"⚠ Warning: Could not add logos: {e}")
