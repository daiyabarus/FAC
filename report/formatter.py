"""Excel formatting utilities"""

from openpyxl.styles import PatternFill, Font, Border, Side, Alignment


class ExcelFormatter:
    """Apply formatting to Excel cells"""

    def __init__(self):
        # PASS / FAIL styles
        self.red_fill = PatternFill(
            start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid"
        )
        self.dark_red_text = Font(color="FF9C0006")

        self.green_fill = PatternFill(
            start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid"
        )
        self.dark_green_text = Font(color="FF006100")

        # Header styles
        self.header_fill = PatternFill("solid", fgColor="47402D")

        self.header_font = Font(bold=True, color="ffffff", size=28)

        # ✅ NEW: smaller header font (same color & style)
        self.header_font_small = Font(bold=True, color="000000", size=16)

        # Border
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Common alignment
        self.center_alignment = Alignment(horizontal="center", vertical="center")

    def format_pass_fail(self, cell, is_pass: bool):
        """Format cell as PASS (green) or FAIL (red)"""
        if is_pass:
            cell.fill = self.green_fill
            cell.font = self.dark_green_text
            cell.value = "PASS"
        else:
            cell.fill = self.red_fill
            cell.font = self.dark_red_text
            cell.value = "FAIL"

        cell.border = self.thin_border
        cell.alignment = self.center_alignment

    def format_header(self, cell, value: str):
        """Format main header cell"""
        cell.fill = self.header_fill
        cell.font = self.header_font
        cell.value = value
        cell.border = self.thin_border
        cell.alignment = self.center_alignment

    def format_header_small(self, cell, value: str):
        """Format smaller header cell (sub-header / category)"""
        cell.fill = self.header_fill
        cell.font = self.header_font_small
        cell.value = value
        cell.border = self.thin_border
        cell.alignment = self.center_alignment

    def format_value(self, cell, value, num_format: str = "0.00"):
        """Format numeric value cell"""
        cell.value = value
        cell.number_format = num_format
        cell.border = self.thin_border
        cell.alignment = self.center_alignment
