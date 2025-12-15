"""
Report Generator Service
Main orchestrator for generating FAC reports with all sheets
"""
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from io import BytesIO
import base64
from datetime import datetime
from services.data_loader import DataLoader
from services.data_processor import DataProcessor
from services.kpi_calculator import KPICalculator
from services.chart_generator import ChartGenerator


class ReportGenerator:
    """Main report generator"""

    TEMPLATE_PATH = "./datatemplate.xlsx"

    def __init__(self):
        self.loader = DataLoader()
        self.processor = DataProcessor()
        self.calculator = KPICalculator()
        self.chart_gen = ChartGenerator()
        self.lte_data = None
        self.gsm_data = None
        self.cluster_data = None
        self.processed_lte = None
        self.processed_gsm = None
        self.kpi_results = None
        self.charts_by_cluster = {}

    def load_data(self, lte_file: str, gsm_file: str, cluster_file: str):
        """Load all data files"""
        self.lte_data = self.loader.load_lte_data(lte_file)
        self.gsm_data = self.loader.load_gsm_data(gsm_file)
        self.cluster_data = self.loader.load_cluster_data(cluster_file)

        validation = self.loader.validate_data()
        if not validation['all_valid']:
            raise Exception(f"Data validation failed: {validation['errors']}")

    def process_data(self):
        """Process and transform data"""
        self.processed_lte, self.processed_gsm = self.processor.process(
            self.lte_data, self.gsm_data, self.cluster_data
        )

    def calculate_kpis(self):
        """Calculate all KPIs"""
        self.kpi_results = self.calculator.calculate_all_kpis(
            self.processed_lte, self.processed_gsm
        )

    def generate_charts(self):
        """Generate charts for all clusters"""
        clusters = self.processor.get_unique_clusters()

        for cluster in clusters:
            charts = self.chart_gen.generate_all_charts(
                self.processed_lte, self.processed_gsm, cluster
            )
            self.charts_by_cluster[cluster] = charts

    def generate_reports(self, output_dir: str) -> list:
        """Generate Excel reports for each cluster"""
        output_files = []
        clusters = self.processor.get_unique_clusters()

        for cluster in clusters:
            output_file = self._generate_cluster_report(cluster, output_dir)
            output_files.append(output_file)

        return output_files

    def _generate_cluster_report(self, cluster: str, output_dir: str) -> str:
        """Generate report for single cluster"""
        # Load template
        wb = openpyxl.load_workbook(self.TEMPLATE_PATH)

        # Rename Template sheet to FAC
        if 'Template' in wb.sheetnames:
            ws_fac = wb['Template']
            ws_fac.title = 'FAC'
        else:
            ws_fac = wb.active
            ws_fac.title = 'FAC'

        # Write KPI results to FAC sheet
        self._write_fac_sheet(ws_fac, cluster)

        # Create Contributors sheet
        ws_contrib = wb.create_sheet('Contributors')
        self._write_contributors_sheet(ws_contrib, cluster)

        # Create RAW 2G sheet
        ws_raw2g = wb.create_sheet('RAW 2G')
        self._write_raw_2g_sheet(ws_raw2g, cluster)

        # Create RAW 4G sheet
        ws_raw4g = wb.create_sheet('RAW 4G')
        self._write_raw_4g_sheet(ws_raw4g, cluster)

        # Create Charts sheet
        ws_charts = wb.create_sheet('Charts')
        self._write_charts_sheet(ws_charts, cluster)

        # Save output
        output_path = Path(
            output_dir) / f"FAC_Report_{cluster}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        wb.save(output_path)

        return str(output_path)

    def _write_fac_sheet(self, ws, cluster):
        """Write KPI results to FAC sheet"""
        # Title in A6
        first_month, last_month = self.processor.get_months_range()
        if first_month and last_month:
            try:
                # Safe date formatting
                start_str = f"{first_month.day} {first_month.strftime('%B')}"
                end_str = f"{last_month.day} {last_month.strftime('%B')}"
                date_range = f"{start_str} to {end_str}"
            except:
                date_range = "Date Range"
        else:
            date_range = "Date Range"

        ws['A6'] = f"FAC KPI Achievement Summary {cluster} - {date_range}"

        # Month headers (M12, O12, Q12)
        months = list(self.kpi_results.get('2g_ran', {}).keys())
        if len(months) > 0:
            ws['M12'] = self._get_month_name(1)
            ws['M13'] = self._get_month_range(1)
        if len(months) > 1:
            ws['O12'] = self._get_month_name(2)
            ws['O13'] = self._get_month_range(2)
        if len(months) > 2:
            ws['Q12'] = self._get_month_name(3)
            ws['Q13'] = self._get_month_range(3)

        # Write 2G KPI results (M14-R16)
        self._write_2g_kpi_results(ws, cluster)

        # Write 4G KPI results (M17-R47)
        self._write_4g_kpi_results(ws, cluster)

        # Overall status (M63, O63, Q63)
        overall = self.calculator.get_overall_status_by_month()
        ws['M63'] = overall.get('month_1', 'N/A')
        ws['O63'] = overall.get('month_2', 'N/A')
        ws['Q63'] = overall.get('month_3', 'N/A')

        # Apply PASS/FAIL formatting
        self._apply_pass_fail_formatting(ws)

    def _write_2g_kpi_results(self, ws, cluster):
        """Write 2G RAN KPI results"""
        cells_map = {
            'Call Setup Success Rate': [('M14', 'N14'), ('O14', 'P14'), ('Q14', 'R14')],
            'SDCCH Success rate': [('M15', 'N15'), ('O15', 'P15'), ('Q15', 'R15')],
            'Perceive Drop Rate': [('M16', 'N16'), ('O16', 'P16'), ('Q16', 'R16')]
        }

        for kpi_name, cell_list in cells_map.items():
            for month_idx, (value_cell, result_cell) in enumerate(cell_list, 1):
                month_key = f'month_{month_idx}'
                if month_key in self.kpi_results.get('2g_ran', {}):
                    kpi_result = self.kpi_results['2g_ran'][month_key].get(
                        kpi_name)
                    if kpi_result:
                        ws[value_cell] = round(kpi_result['value'], 2)
                        ws[result_cell] = kpi_result['status']

    def _write_4g_kpi_results(self, ws, cluster):
        """Write 4G RAN KPI results"""
        # Define cell mappings for each KPI
        simple_kpis = {
            'Session Setup Success Rate': [('M17', 'N17'), ('O17', 'P17'), ('Q17', 'R17')],
            'Handover Success Rate Inter and Intra-Frequency': [('M20', 'N20'), ('O20', 'P20'), ('Q20', 'R20')],
            'E-RAB Drop Rate': [('M21', 'N21'), ('O21', 'P21'), ('Q21', 'R21')],
            'UL Packet Loss (PDCP )': [('M26', 'N26'), ('O26', 'P26'), ('Q26', 'R26')],
            'DL Packet Loss (PDCP )': [('M27', 'N27'), ('O27', 'P27'), ('Q27', 'R27')],
            'CQI': [('M28', 'N28'), ('O28', 'P28'), ('Q28', 'R28')],
            'UL RSSI': [('M31', 'N31'), ('O31', 'P31'), ('Q31', 'R31')],
            'LTC Non Capacity': [('M34', 'N34'), ('O34', 'P34'), ('Q34', 'R34')],
            'Coverage Overlaping ratio ': [('M35', 'N35'), ('O35', 'P35'), ('Q35', 'R35')],
            'Voice Call Success Rate (VoLTE)': [('M45', 'N45'), ('O45', 'P45'), ('Q45', 'R45')],
            'Voice Call Drop Rate (VoLTE)': [('M46', 'N46'), ('O46', 'P46'), ('Q46', 'R46')],
            'SRVCC Success Rate': [('M47', 'N47'), ('O47', 'P47'), ('Q47', 'R47')]
        }

        # Write simple KPIs
        for kpi_name, cell_list in simple_kpis.items():
            for month_idx, (value_cell, result_cell) in enumerate(cell_list, 1):
                month_key = f'month_{month_idx}'
                if month_key in self.kpi_results.get('4g_ran', {}):
                    kpi_result = self.kpi_results['4g_ran'][month_key].get(
                        kpi_name)
                    if isinstance(kpi_result, dict):
                        ws[value_cell] = round(kpi_result['value'], 2)
                        ws[result_cell] = kpi_result['status']

        # Multi-condition KPIs
        multi_kpis = {
            'RACH Success Rate': [
                [('M18', 'N18'), ('O18', 'P18'), ('Q18', 'R18')],
                [('M19', 'N19'), ('O19', 'P19'), ('Q19', 'R19')]
            ],
            'Downlink User Throughput': [
                [('M22', 'N22'), ('O22', 'P22'), ('Q22', 'R22')],
                [('M23', 'N23'), ('O23', 'P23'), ('Q23', 'R23')]
            ],
            'Uplink User Throughput': [
                [('M24', 'N24'), ('O24', 'P24'), ('Q24', 'R24')],
                [('M25', 'N25'), ('O25', 'P25'), ('Q25', 'R25')]
            ],
            'MIMO Transmission Rank2 Rate': [
                [('M29', 'N29'), ('O29', 'P29'), ('Q29', 'R29')],
                [('M30', 'N30'), ('O30', 'P30'), ('Q30', 'R30')]
            ],
            'Packet Latency': [
                [('M32', 'N32'), ('O32', 'P32'), ('Q32', 'R32')],
                [('M33', 'N33'), ('O33', 'P33'), ('Q33', 'R33')]
            ]
        }

        for kpi_name, conditions_cells in multi_kpis.items():
            for month_idx in range(1, 4):
                month_key = f'month_{month_idx}'
                if month_key in self.kpi_results.get('4g_ran', {}):
                    kpi_results = self.kpi_results['4g_ran'][month_key].get(
                        kpi_name, [])
                    if isinstance(kpi_results, list):
                        for cond_idx, result in enumerate(kpi_results):
                            if cond_idx < len(conditions_cells):
                                value_cell, result_cell = conditions_cells[cond_idx][month_idx - 1]
                                ws[value_cell] = round(result['value'], 2)
                                ws[result_cell] = result['status']

        # Spectral Efficiency (dynamic)
        se_cells = {
            'month_1': [
                ('M36', 'N36'), ('M37', 'N37'), ('M38', 'N38'), ('M39', 'N39'),
                ('M40', 'N40'), ('M41', 'N41'), ('M42', 'N42')
            ],
            'month_2': [
                ('O36', 'P36'), ('O37', 'P37'), ('O38', 'P38'), ('O39', 'P39'),
                ('O40', 'P40'), ('O41', 'P41'), ('O42', 'P42')
            ],
            'month_3': [
                ('Q36', 'R36'), ('Q37', 'R37'), ('Q38', 'R38'), ('Q39', 'R39'),
                ('Q40', 'R40'), ('Q41', 'R41'), ('Q42', 'R42')
            ]
        }

        for month_idx in range(1, 4):
            month_key = f'month_{month_idx}'
            if month_key in self.kpi_results.get('4g_ran', {}):
                se_results = self.kpi_results['4g_ran'][month_key].get(
                    'Spectral Efficiency', [])
                if isinstance(se_results, list):
                    for se_idx, result in enumerate(se_results):
                        if se_idx < len(se_cells[month_key]):
                            value_cell, result_cell = se_cells[month_key][se_idx]
                            ws[value_cell] = round(result['value'], 2)
                            ws[result_cell] = result['status']

    def _write_contributors_sheet(self, ws, cluster):
        """Write contributors (failed cells) to Contributors sheet"""
        headers = ['Month', 'Clause Type', 'KPI Domain', 'Name', 'Reference',
                   'TOWER_ID', 'CELL_NAME', 'Value', 'Baseline', 'Status']

        ws.append(headers)

        # Filter contributors for this cluster
        cluster_lte = self.processed_lte[self.processed_lte['CLUSTER'] == cluster]
        cluster_gsm = self.processed_gsm[self.processed_gsm['CLUSTER'] == cluster]

        for contrib in self.kpi_results.get('contributors', []):
            if contrib.get('TOWER_ID') in cluster_lte['TOWER_ID'].values or \
               contrib.get('TOWER_ID') in cluster_gsm['TOWER_ID'].values:
                ws.append([
                    contrib.get('Month'),
                    contrib.get('Clause Type'),
                    contrib.get('KPI Domain'),
                    contrib.get('Name'),
                    contrib.get('Reference'),
                    contrib.get('TOWER_ID'),
                    contrib.get('CELL_NAME'),
                    contrib.get('Value'),
                    contrib.get('Baseline'),
                    contrib.get('Status')
                ])

    def _write_raw_2g_sheet(self, ws, cluster):
        """Write RAW 2G data"""
        gsm_cluster = self.processed_gsm[self.processed_gsm['CLUSTER'] == cluster].copy(
        )

        # Convert Period columns to string for Excel compatibility
        if 'YEAR_MONTH' in gsm_cluster.columns:
            gsm_cluster['YEAR_MONTH'] = gsm_cluster['YEAR_MONTH'].astype(str)

        # Convert any remaining Period types
        for col in gsm_cluster.columns:
            if gsm_cluster[col].dtype.name == 'period[M]':
                gsm_cluster[col] = gsm_cluster[col].astype(str)

        for r in dataframe_to_rows(gsm_cluster, index=False, header=True):
            ws.append(r)

    def _write_raw_4g_sheet(self, ws, cluster):
        """Write RAW 4G data"""
        lte_cluster = self.processed_lte[self.processed_lte['CLUSTER'] == cluster].copy(
        )

        # Convert Period columns to string for Excel compatibility
        if 'YEAR_MONTH' in lte_cluster.columns:
            lte_cluster['YEAR_MONTH'] = lte_cluster['YEAR_MONTH'].astype(str)

        # Convert any remaining Period types
        for col in lte_cluster.columns:
            if lte_cluster[col].dtype.name == 'period[M]':
                lte_cluster[col] = lte_cluster[col].astype(str)

        for r in dataframe_to_rows(lte_cluster, index=False, header=True):
            ws.append(r)

    def _write_charts_sheet(self, ws, cluster):
        """Write charts to Charts sheet"""
        charts = self.charts_by_cluster.get(cluster, [])

        row = 1
        for chart_data in charts:
            img_data = chart_data['data']
            img = Image(img_data)
            img.width = 800
            img.height = 400

            ws.add_image(img, f'A{row}')
            row += 25  # Space between charts

    def _apply_pass_fail_formatting(self, ws):
        """Apply color formatting to PASS/FAIL cells"""
        pass_fill = PatternFill(start_color="00FF00",
                                end_color="00FF00", fill_type="solid")
        fail_fill = PatternFill(start_color="FF0000",
                                end_color="FF0000", fill_type="solid")
        pass_font = Font(color="FFFFFF", bold=True)
        fail_font = Font(color="FFFFFF", bold=True)

        # Check result columns (N, P, R from rows 14-47 and 63)
        result_cols = ['N', 'P', 'R']

        for col in result_cols:
            for row in range(14, 48):
                cell = ws[f'{col}{row}']
                if cell.value == 'PASS':
                    cell.fill = pass_fill
                    cell.font = pass_font
                elif cell.value == 'FAIL':
                    cell.fill = fail_fill
                    cell.font = fail_font

            # Overall status row
            cell = ws[f'{col}63']
            if cell.value == 'PASS':
                cell.fill = pass_fill
                cell.font = pass_font
            elif cell.value == 'FAIL':
                cell.fill = fail_fill
                cell.font = fail_font

    def _get_month_name(self, month_num: int) -> str:
        """Get month name for display"""
        if self.processed_lte is not None and 'MONTH_NAME' in self.processed_lte.columns:
            months = sorted(self.processed_lte['MONTH_NAME'].unique())
            if len(months) >= month_num:
                return months[month_num - 1]
        return f"Month {month_num}"

    def _get_month_range(self, month_num: int) -> str:
        """Get month date range for display"""
        try:
            if self.processed_lte is not None and 'YEAR_MONTH' in self.processed_lte.columns:
                months = sorted(
                    self.processed_lte['YEAR_MONTH'].dropna().unique())
                if len(months) >= month_num:
                    month_period = months[month_num - 1]
                    start = month_period.start_time
                    end = month_period.end_time

                    # Safe formatting - handle both Windows and Unix
                    try:
                        # Try Unix format first (%-d)
                        return f"{start.strftime('%-d %B')} - {end.strftime('%-d %B')}"
                    except:
                        # Fallback to Windows format (%#d on Windows, %d elsewhere)
                        start_day = start.day
                        end_day = end.day
                        start_month = start.strftime('%B')
                        end_month = end.strftime('%B')
                        return f"{start_day} {start_month} - {end_day} {end_month}"
        except Exception as e:
            print(f"Warning: Error formatting date range: {e}")

        return f"Date Range Month {month_num}"
